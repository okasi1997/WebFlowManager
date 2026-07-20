"""PCL のスキーマ編集、階層データ編集、JSON/Excel 入出力を扱う。"""
from __future__ import annotations
import copy
import json
import re
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any
from core.database import Database
from i18n import SUPPORTED_LANGUAGES, tr, tr_language
from ui.ui_helpers import scrollable_tree
TYPES = ('text', 'number', 'boolean', 'object', 'list')

def validate_schema(node: Any, location: str='Data') -> None:
    if not isinstance(node, dict) or not isinstance(node.get('name'), str) or (not node['name'].strip()):
        raise ValueError(f'{location}msg.0236')
    if node.get('type') not in TYPES:
        raise ValueError(f'{location}msg.0237{node.get('type')}')
    children = node.get('children', [])
    if node['type'] in ('object', 'list'):
        if not isinstance(children, list):
            raise ValueError(f'{location}msg.0238')
        names: set[str] = set()
        for child in children:
            child_name = child.get('name') if isinstance(child, dict) else '?'
            if child_name in names:
                raise ValueError(f'{location}msg.0239{child_name}')
            names.add(child_name)
            validate_schema(child, f'{location}.{child_name}')
    elif 'children' in node and children:
        raise ValueError(f'{location}msg.0240')

def scalar_paths(schema: dict[str, Any]) -> list[str]:
    return [path for path, node in schema_paths(schema) if node['type'] not in ('object', 'list')]

def flatten_record(schema: dict[str, Any], data: dict[str, Any]) -> list[dict[str, Any]]:
    """入れ子の list を、Excel に書き出せる葉の行へ展開する。"""
    return [values for values, _groups in _flatten_record_with_groups(schema, data)]

def _flatten_record_with_groups(schema: dict[str, Any], data: dict[str, Any]) -> list[tuple[dict[str, Any], dict[str, int]]]:
    # list の添字も返し、同じ親値を Excel 上で何度も出力しないようにする。

    def merge(left: list[tuple[dict[str, Any], dict[str, int]]], right: list[tuple[dict[str, Any], dict[str, int]]]) -> list[tuple[dict[str, Any], dict[str, int]]]:
        return [({**a_values, **b_values}, {**a_groups, **b_groups}) for a_values, a_groups in left for b_values, b_groups in right]

    def walk_fields(nodes: list[dict[str, Any]], value: dict[str, Any], prefix: str) -> list[tuple[dict[str, Any], dict[str, int]]]:
        rows: list[tuple[dict[str, Any], dict[str, int]]] = [({}, {})]
        for node in nodes:
            path = f'{prefix}.{node['name']}' if prefix else node['name']
            current = value.get(node['name'])
            if node['type'] == 'list':
                items = current if isinstance(current, list) else []
                child_rows: list[tuple[dict[str, Any], dict[str, int]]] = []
                for item_index, item in enumerate(items):
                    for child_values, child_groups in walk_fields(node.get('children', []), item if isinstance(item, dict) else {}, path):
                        child_rows.append((child_values, {path: item_index, **child_groups}))
                if not child_rows:
                    child_rows = [({child_path: '' for child_path in _descendant_scalar_paths(node, path)}, {path: -1})]
            elif node['type'] == 'object':
                child_rows = walk_fields(node.get('children', []), current if isinstance(current, dict) else {}, path)
            else:
                child_rows = [({path: current if current is not None else ''}, {})]
            rows = merge(rows, child_rows)
        return rows
    return walk_fields(schema.get('children', []), data, '')

def scalar_list_owners(schema: dict[str, Any]) -> dict[str, tuple[str, ...]]:
    result: dict[str, tuple[str, ...]] = {}

    def walk(node: dict[str, Any], prefix: str, list_ancestors: tuple[str, ...]) -> None:
        for child in node.get('children', []):
            path = f'{prefix}.{child['name']}' if prefix else child['name']
            ancestors = (*list_ancestors, path) if child['type'] == 'list' else list_ancestors
            if child['type'] in ('object', 'list'):
                walk(child, path, ancestors)
            else:
                result[path] = list_ancestors
    walk(schema, '', ())
    return result

def _descendant_scalar_paths(node: dict[str, Any], prefix: str) -> list[str]:
    result: list[str] = []
    for child in node.get('children', []):
        path = f'{prefix}.{child['name']}'
        if child['type'] in ('object', 'list'):
            result.extend(_descendant_scalar_paths(child, path))
        else:
            result.append(path)
    return result

def write_records_excel(path: str | Path, schema: dict[str, Any], records: list[dict[str, Any]]) -> int:
    """PCL と実行設定を、再読込可能な Excel ブックへ保存する。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    columns = scalar_paths(schema)
    owners = scalar_list_owners(schema)
    max_depth = max((len(column.split('.')) for column in columns), default=1)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = tr('msg.0241')
    sheet.cell(1, 1, tr('msg.0242'))
    if max_depth > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=max_depth, end_column=1)
    path_parts = [column.split('.') for column in columns]
    for column_index, parts in enumerate(path_parts, 2):
        for level, part in enumerate(parts, 1):
            sheet.cell(level, column_index, part)
        if len(parts) < max_depth:
            sheet.merge_cells(start_row=len(parts), start_column=column_index, end_row=max_depth, end_column=column_index)
    for level in range(1, max_depth + 1):
        start = 0
        while start < len(columns):
            prefix = tuple(path_parts[start][:level]) if len(path_parts[start]) >= level else None
            end = start + 1
            while end < len(columns) and prefix is not None and (tuple(path_parts[end][:level]) == prefix):
                end += 1
            if prefix is not None and end - start > 1:
                sheet.merge_cells(start_row=level, start_column=start + 2, end_row=level, end_column=end + 1)
            start = end
    output_row = max_depth + 1
    for record in records:
        rows = _flatten_record_with_groups(schema, normalize_record(schema, record['data'])) or [({}, {})]
        seen_scopes: set[tuple[Any, ...]] = set()
        for row_index, (values, groups) in enumerate(rows):
            sheet.cell(output_row, 1, record['name'] if row_index == 0 else '')
            for column_index, column in enumerate(columns, 2):
                ancestors = owners[column]
                scope = (column, *(groups.get(path, -1) for path in ancestors))
                if scope not in seen_scopes:
                    sheet.cell(output_row, column_index, values.get(column, ''))
                    seen_scopes.add(scope)
            output_row += 1
    for row in sheet.iter_rows(min_row=1, max_row=max_depth):
        for cell in row:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')
    sheet.freeze_panes = f'A{max_depth + 1}'
    headers = [tr('msg.0242'), *columns]
    for index, header in enumerate(headers, 1):
        values = [str(sheet.cell(row, index).value or '') for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, *(len(value) + 2 for value in values)), 45)
    settings_sheet = workbook.create_sheet(tr('msg.0243'))
    settings_sheet.append([tr('msg.0242'), tr('msg.0244'), tr('msg.0245')])
    for record in records:
        settings_sheet.append([record['name'], bool(record.get('enabled', True)), str(record.get('execution_group', '1'))])
    workbook.save(path)
    return sheet.max_row - max_depth

def read_records_excel(path: str | Path, schema: dict[str, Any]) -> list[dict[str, Any]]:
    """write_records_excel が作成した階層ブックを PCL に復元する。"""
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    header_depth = 1
    for merged in sheet.merged_cells.ranges:
        # 先頭列の縦結合範囲が、階層ヘッダーの深さを表す。
        if merged.min_col == merged.max_col == 1 and merged.min_row == 1:
            header_depth = max(header_depth, merged.max_row)

    def header_value(row: int, column: int) -> Any:
        value = sheet.cell(row, column).value
        if value is not None:
            return value
        for merged in sheet.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
                return sheet.cell(merged.min_row, merged.min_col).value
        return None
    columns: list[str] = []
    for column in range(2, sheet.max_column + 1):
        parts = [str(header_value(row, column)).strip() for row in range(1, header_depth + 1) if header_value(row, column) not in (None, '')]
        path_name = '.'.join(dict.fromkeys(parts))
        columns.append(path_name)
    expected = set(scalar_paths(schema))
    if not columns or any((column not in expected for column in columns)):
        raise ValueError('msg.0246')
    grouped: list[tuple[str, list[dict[str, Any]]]] = []
    current_name = ''
    current_rows: list[dict[str, Any]] = []
    for row_number in range(header_depth + 1, sheet.max_row + 1):
        name_value = sheet.cell(row_number, 1).value
        values = {path_name: sheet.cell(row_number, index + 2).value for index, path_name in enumerate(columns)}
        if name_value not in (None, ''):
            if current_name:
                grouped.append((current_name, current_rows))
            current_name = str(name_value).strip()
            current_rows = []
        if current_name and any((value is not None for value in values.values())):
            current_rows.append(values)
    if current_name:
        grouped.append((current_name, current_rows))
    if not grouped:
        raise ValueError('msg.0247')
    execution_settings: list[tuple[bool, str]] = []
    settings_names = {tr_language('msg.0243', language) for language in SUPPORTED_LANGUAGES}
    # 出力時と現在の UI 言語が異なっても設定シートを認識する。
    settings_name = next((name for name in workbook.sheetnames if name in settings_names), None)
    if settings_name is not None:
        settings_sheet = workbook[settings_name]
        for row in settings_sheet.iter_rows(min_row=2, values_only=True):
            enabled_value = row[1] if len(row) > 1 else True
            enabled_words = {tr_language('msg.0037', language).lower() for language in SUPPORTED_LANGUAGES}
            enabled_words.add(tr('msg.0248').lower())
            enabled = enabled_value if isinstance(enabled_value, bool) else str(enabled_value).lower() in {'true', '1', 'yes', *enabled_words}
            group = str(row[2] if len(row) > 2 and row[2] not in (None, '') else '1').strip()
            execution_settings.append((enabled, group))
    owners = scalar_list_owners(schema)
    list_paths = sorted({owner for value in owners.values() for owner in value}, key=lambda value: (value.count('.'), value))
    anchors = {list_path: [column for column in columns if owners[column] and owners[column][-1] == list_path] for list_path in list_paths}
    node_by_path = {path_name: node for path_name, node in schema_paths(schema)}

    def assign(data: dict[str, Any], path_name: str, value: Any, indexes: dict[str, int]) -> None:
        # パスと行ごとの list 添字を使って、平坦なセルを階層辞書へ戻す。
        container: Any = data
        prefix = ''
        parts = path_name.split('.')
        for index, part in enumerate(parts):
            prefix = f'{prefix}.{part}' if prefix else part
            node = node_by_path[prefix]
            last = index == len(parts) - 1
            if node['type'] == 'list':
                target = container.setdefault(part, [])
                item_index = indexes[prefix]
                while len(target) <= item_index:
                    target.append(new_list_item(node))
                container = target[item_index]
            elif last:
                if node['type'] == 'number' and value not in (None, ''):
                    value = float(value) if isinstance(value, float) and (not value.is_integer()) else int(value)
                elif node['type'] == 'boolean':
                    yes_words = {tr_language('msg.0037', language).lower() for language in SUPPORTED_LANGUAGES}
                    value = value if isinstance(value, bool) else str(value).lower() in {'true', '1', 'yes', *yes_words}
                container[part] = value
            elif node['type'] == 'object':
                container = container.setdefault(part, {})
    result: list[dict[str, Any]] = []
    for name, rows in grouped:
        data = {child['name']: default_value(child) for child in schema.get('children', [])}
        counters = {path_name: 0 for path_name in list_paths}
        previous = dict(counters)
        for row_index, values in enumerate(rows):
            previous = dict(counters)
            for list_path in list_paths:
                parents = [owner for owner in list_paths if list_path.startswith(owner + '.')]
                parent = max(parents, key=len) if parents else None
                if row_index == 0 or (parent and counters[parent] != previous[parent]):
                    counters[list_path] = 0
                elif any((values.get(column) is not None for column in anchors[list_path])):
                    counters[list_path] += 1
            for path_name, value in values.items():
                if value is not None:
                    assign(data, path_name, value, counters)
        setting = execution_settings[len(result)] if len(result) < len(execution_settings) else (True, '1')
        result.append({'name': name, 'enabled': setting[0], 'execution_group': setting[1], 'data': normalize_record(schema, data)})
    return result

def default_value(node: dict[str, Any]) -> Any:
    kind = node['type']
    if kind in ('object', 'list'):
        value = {child['name']: default_value(child) for child in node.get('children', [])}
        return [] if kind == 'list' else value
    if kind == 'number':
        return 0
    if kind == 'boolean':
        return False
    return ''

def new_list_item(node: dict[str, Any], populate_nested_lists: bool=False) -> dict[str, Any]:

    def build(child: dict[str, Any]) -> Any:
        if child['type'] == 'list':
            return [new_list_item(child, True)] if populate_nested_lists else []
        if child['type'] == 'object':
            return {grandchild['name']: build(grandchild) for grandchild in child.get('children', [])}
        return default_value(child)
    return {child['name']: build(child) for child in node.get('children', [])}

def normalize_record(schema: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    result = dict(data)
    for child in schema.get('children', []):
        name = child['name']
        if name not in result:
            result[name] = default_value(child)
        elif child['type'] == 'object':
            result[name] = normalize_record(child, result[name]) if isinstance(result[name], dict) else default_value(child)
        elif child['type'] == 'list':
            result[name] = [normalize_record(child, item) if isinstance(item, dict) else new_list_item(child) for item in result[name]] if isinstance(result[name], list) else []
    return result

def schema_paths(schema: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    result: list[tuple[str, dict[str, Any]]] = []

    def walk(node: dict[str, Any], prefix: str) -> None:
        for child in node.get('children', []):
            path = f'{prefix}.{child['name']}' if prefix else child['name']
            result.append((path, child))
            if child['type'] in ('object', 'list'):
                walk(child, path)
    walk(schema, '')
    return result

class FieldDialog(tk.Toplevel):

    def __init__(self, parent: tk.Misc, node: dict[str, Any] | None=None) -> None:
        super().__init__(parent)
        self.title('msg.0249')
        self.geometry('460x210')
        self.resizable(False, False)
        self.result: dict[str, Any] | None = None
        self.name = tk.StringVar(value=(node or {}).get('name', ''))
        self.kind = tk.StringVar(value=(node or {}).get('type', 'text'))
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)
        ttk.Separator(self, style='DialogFooter.TSeparator').grid(row=0, column=0, sticky='ew')
        # Optical centering: the long controls carry more visual weight than
        # the labels, so shift the form slightly left while keeping the footer
        # actions mathematically centered.
        main = ttk.Frame(self, padding=(28, 14, 68, 10))
        main.grid(row=1, column=0, sticky='nsew')
        body = ttk.Frame(main)
        body.pack(fill='x', expand=True)
        body.columnconfigure(1, weight=1)
        body.columnconfigure(0, minsize=72)
        ttk.Label(body, text='msg.0250').grid(row=0, column=0, padx=(0, 12), pady=(4, 7), sticky='e')
        ttk.Entry(body, textvariable=self.name, style='Dialog.TEntry').grid(row=0, column=1, pady=(4, 7), sticky='ew')
        ttk.Label(body, text='msg.0251').grid(row=1, column=0, padx=(0, 12), pady=(7, 4), sticky='e')
        ttk.Combobox(body, textvariable=self.kind, values=TYPES, state='readonly', style='Dialog.TCombobox').grid(row=1, column=1, pady=(7, 4), sticky='ew')
        ttk.Separator(self, style='DialogFooter.TSeparator').grid(row=2, column=0, sticky='ew')
        buttons = ttk.Frame(self, style='DialogFooter.TFrame', padding=(14, 10))
        buttons.grid(row=3, column=0, sticky='ew')
        button_group = ttk.Frame(buttons, style='DialogFooter.TFrame')
        button_group.pack(side='right')
        ttk.Button(button_group, text='msg.0144', command=self._save, style='Primary.TButton', width=14).grid(row=0, column=0, padx=(0, 5), sticky='ew')
        ttk.Button(button_group, text='msg.0145', command=self.destroy, style='Secondary.TButton', width=14).grid(row=0, column=1, padx=(5, 0), sticky='ew')
        button_group.columnconfigure(0, weight=1, uniform='field_footer_action')
        button_group.columnconfigure(1, weight=1, uniform='field_footer_action')
        self.transient(parent)
        self.grab_set()
        self.protocol('WM_DELETE_WINDOW', self.destroy)

    def _save(self) -> None:
        name = self.name.get().strip()
        if not name or '.' in name:
            messagebox.showerror('msg.0253', 'msg.0254', parent=self)
            return
        self.result = {'name': name, 'type': self.kind.get()}
        if self.kind.get() in ('object', 'list'):
            self.result['children'] = []
        self.destroy()

class SchemaDesignerDialog(tk.Toplevel):
    """PCL 全体で共有するデータ構造を編集する。"""

    def __init__(self, parent: tk.Misc, db: Database, workflow_id: int, workflow_name: str) -> None:
        super().__init__(parent)
        self.db, self.workflow_id = (db, workflow_id)
        self.schema = copy.deepcopy(db.get_data_schema(workflow_id))
        self.node_by_item: dict[str, tuple[dict[str, Any], dict[str, Any] | None]] = {}
        self.title(f'msg.0255{workflow_name}')
        self.geometry('720x560')
        ttk.Label(self, text='msg.0256', style='Section.TLabel').pack(anchor='w', padx=10, pady=10)
        tree_frame, self.tree = scrollable_tree(self, columns=('type', 'path'), show='tree headings')
        self.tree.heading('#0', text='msg.0250')
        self.tree.heading('type', text='msg.0257')
        self.tree.heading('path', text='msg.0258')
        self.tree.column('#0', width=180)
        self.tree.column('type', width=90)
        self.tree.column('path', width=260)
        tree_frame.pack(fill='both', expand=True, padx=10)
        buttons = ttk.Frame(self)
        buttons.pack(fill='x', padx=10, pady=10)
        button_specs = (('msg.0259', self._add, 'Action.TButton'), ('msg.0260', self._edit, 'TButton'), ('msg.0261', self._delete, 'Danger.TButton'), ('msg.0011', lambda: self._move(-1), 'TButton'), ('msg.0012', lambda: self._move(1), 'TButton'), ('msg.0262', self._export_json, 'TButton'), ('msg.0263', self._import_json, 'TButton'), ('msg.0264', self._save, 'Primary.TButton'))
        action_buttons: list[ttk.Button] = []
        for index, (text, command, button_style) in enumerate(button_specs):
            button = ttk.Button(buttons, text=text, command=command, style=button_style)
            button.grid(row=index // 4, column=index % 4, padx=3, pady=3, sticky='ew')
            action_buttons.append(button)
        self.root_locked_buttons = tuple(action_buttons[index] for index in (1, 2, 3, 4))
        for column in range(4):
            buttons.columnconfigure(column, weight=1)
        self.tree.tag_configure('schema_root', background='#F3F3F3', foreground='#737373', font=(self.db.get_ui_font()[0], self.db.get_ui_font()[1], 'bold'))
        self.tree.bind('<<TreeviewSelect>>', self._update_action_buttons)
        self._refresh()
        self.transient(parent)

    def _refresh(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.node_by_item.clear()
        # The first row is the schema container, not an editable business
        # field.  Keep the persisted schema name for compatibility, while
        # presenting a neutral UI label.
        root = self.tree.insert('', 'end', text='Data', values=('list', ''), open=True, tags=('schema_root',))
        self.node_by_item[root] = (self.schema, None)

        def add(parent_item: str, parent_node: dict[str, Any], prefix: str) -> None:
            for node in parent_node.get('children', []):
                path = f'{prefix}.{node['name']}' if prefix else node['name']
                item = self.tree.insert(parent_item, 'end', text=node['name'], values=(node['type'], path), open=True)
                self.node_by_item[item] = (node, parent_node)
                add(item, node, path)
        add(root, self.schema, '')
        self._update_action_buttons()

    def _update_action_buttons(self, _event: object=None) -> None:
        selected = self._selected()
        editable = selected is not None and selected[1] is not None
        state = 'normal' if editable else 'disabled'
        for button in self.root_locked_buttons:
            button.configure(state=state)

    def _selected(self) -> tuple[dict[str, Any], dict[str, Any] | None] | None:
        selected = self.tree.selection()
        return self.node_by_item.get(selected[0]) if selected else None

    def _add(self) -> None:
        selected = self._selected()
        parent = selected[0] if selected and selected[0]['type'] in ('object', 'list') else selected[1] if selected else self.schema
        if parent is None:
            return
        dialog = FieldDialog(self)
        self.wait_window(dialog)
        if not dialog.result:
            return
        if any((child['name'] == dialog.result['name'] for child in parent.get('children', []))):
            messagebox.showerror('msg.0265', 'msg.0266', parent=self)
            return
        parent.setdefault('children', []).append(dialog.result)
        self._refresh()

    def _edit(self) -> None:
        selected = self._selected()
        if not selected or selected[1] is None:
            return
        node, parent = selected
        dialog = FieldDialog(self, node)
        self.wait_window(dialog)
        if not dialog.result:
            return
        if any((child is not node and child['name'] == dialog.result['name'] for child in parent['children'])):
            messagebox.showerror('msg.0265', 'msg.0266', parent=self)
            return
        children = node.get('children', []) if dialog.result['type'] in ('object', 'list') else None
        node.clear()
        node.update(dialog.result)
        if children is not None:
            node['children'] = children
        self._refresh()

    def _delete(self) -> None:
        selected = self._selected()
        if selected and selected[1] is not None and messagebox.askyesno('msg.0046', 'msg.0267', parent=self):
            selected[1]['children'].remove(selected[0])
            self._refresh()

    def _move(self, direction: int) -> None:
        selected = self._selected()
        if not selected or selected[1] is None:
            return
        children = selected[1]['children']
        index = children.index(selected[0])
        target = index + direction
        if 0 <= target < len(children):
            children[index], children[target] = (children[target], children[index])
            self._refresh()

    def _save(self) -> None:
        try:
            validate_schema(self.schema)
        except ValueError as error:
            messagebox.showerror('msg.0268', str(error), parent=self)
            return
        self.db.save_data_schema(self.workflow_id, self.schema)
        for record in self.db.list_data_records(self.workflow_id):
            synchronized = normalize_record(self.schema, record['data'])
            self.db.update_data_record(record['id'], record['name'], synchronized)
        self.destroy()

    def _export_json(self) -> None:
        path = filedialog.asksaveasfilename(parent=self, defaultextension='.json', filetypes=(('Json', '*.json'),))
        if not path:
            return
        payload = {'version': 1, 'type': 'web-flow-schema', 'schema': self.schema}
        Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        messagebox.showinfo('msg.0269', f'msg.0270{path}', parent=self)

    def _import_json(self) -> None:
        path = filedialog.askopenfilename(parent=self, filetypes=(('Json', '*.json'), ('msg.0054', '*.*')))
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding='utf-8-sig'))
            imported = payload.get('schema') if isinstance(payload, dict) and 'schema' in payload else payload
            validate_schema(imported)
            if imported['type'] != 'list':
                raise ValueError('msg.0271')
        except (OSError, json.JSONDecodeError, ValueError) as error:
            messagebox.showerror('msg.0057', str(error), parent=self)
            return
        if not messagebox.askyesno('msg.0272', 'msg.0273', parent=self):
            return
        self.schema = copy.deepcopy(imported)
        self.db.save_data_schema(self.workflow_id, self.schema)
        for record in self.db.list_data_records(self.workflow_id):
            self.db.update_data_record(record['id'], record['name'], normalize_record(self.schema, record['data']))
        self._refresh()
        messagebox.showinfo('msg.0274', 'msg.0275', parent=self)

class DataPathDialog(tk.Toplevel):

    def __init__(self, parent: tk.Misc, schema: dict[str, Any], lists_only: bool=False) -> None:
        super().__init__(parent)
        self.title('msg.0276')
        self.geometry('520x460')
        self.result: str | None = None
        self.paths: dict[str, tuple[str, str]] = {}
        self.lists_only = lists_only
        tree_frame, self.tree = scrollable_tree(self, columns=('type', 'path'), show='tree headings')
        self.tree.heading('#0', text='msg.0277')
        self.tree.heading('type', text='msg.0257')
        self.tree.heading('path', text='msg.0278')
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        root = self.tree.insert('', 'end', text='msg.0279', values=('object', ''), open=True)
        parent_items: dict[str, str] = {'': root}
        for path, node in schema_paths(schema):
            parent_path = path.rpartition('.')[0]
            item = self.tree.insert(parent_items[parent_path], 'end', text=node['name'], values=(node['type'], path), open=True)
            parent_items[path] = item
            self.paths[item] = (path, node['type'])
        ttk.Button(self, text='msg.0280', command=self._choose).pack(pady=(0, 10))
        self.tree.bind('<Double-1>', lambda _event: self._choose())
        self.transient(parent)
        self.grab_set()

    def _choose(self) -> None:
        selection = self.tree.selection()
        if not selection or selection[0] not in self.paths:
            return
        path, kind = self.paths[selection[0]]
        if self.lists_only and kind != 'list':
            messagebox.showinfo('msg.0281', 'msg.0282', parent=self)
            return
        if not self.lists_only and kind in ('list', 'object'):
            messagebox.showinfo('msg.0283', 'msg.0284', parent=self)
            return
        self.result = path
        self.destroy()

class HierarchicalDataDialog(tk.Toplevel):
    """PCL レコードと、その入れ子データおよび実行設定を編集する。"""

    def __init__(self, parent: tk.Misc, db: Database, workflow_id: int, workflow_name: str) -> None:
        super().__init__(parent)
        self.db, self.workflow_id = (db, workflow_id)
        self.schema = db.get_data_schema(workflow_id)
        self.current_id: int | None = None
        self.current_name = ''
        self.current_data: dict[str, Any] = {}
        self.meta: dict[str, dict[str, Any]] = {}
        self.view_identity: dict[str, str] = {}
        self.force_open_identities: set[str] = set()
        self.force_select_identity: str | None = None
        self.record_sort_column: str | None = None
        self.record_sort_descending = False
        self.title(f'msg.0285{workflow_name}')
        self.geometry('1050x650')
        pane = ttk.Panedwindow(self, orient='horizontal')
        pane.pack(fill='both', expand=True, padx=10, pady=10)
        left, right = (ttk.Frame(pane), ttk.Frame(pane))
        pane.add(left, weight=1)
        pane.add(right, weight=3)
        pcl_header = ttk.Frame(left)
        pcl_header.pack(fill='x')
        ttk.Label(pcl_header, text='msg.0286', style='Section.TLabel').pack(side='left')
        self.session_limit = tk.StringVar(value=str(self.db.get_pcl_session_limit()))
        session_spinbox = ttk.Spinbox(pcl_header, from_=1, to=20, width=4, textvariable=self.session_limit, command=self._save_session_limit)
        session_spinbox.pack(side='right')
        ttk.Label(pcl_header, text='msg.0287').pack(side='right', padx=(12, 2))
        session_spinbox.bind('<Return>', self._save_session_limit)
        session_spinbox.bind('<FocusOut>', self._save_session_limit)
        records_frame, self.records = scrollable_tree(left, columns=('name', 'enabled', 'group'), show='headings')
        self.record_heading_keys = {'name': 'msg.0218', 'enabled': 'msg.0244', 'group': 'msg.0245'}
        self._update_record_headings()
        self.records.column('name', width=140, minwidth=100, stretch=True)
        self.records.column('enabled', width=90, minwidth=80, anchor='center', stretch=False)
        self.records.column('group', width=110, minwidth=100, anchor='center', stretch=False)
        self.records.tag_configure('disabled', foreground='#A0A0A0')
        records_frame.pack(fill='both', expand=True, pady=5)
        self.records.bind('<<TreeviewSelect>>', self._select_record)
        self.records.bind('<Double-1>', self._record_double_click)
        lb = ttk.Frame(left)
        lb.pack(fill='x')
        record_buttons = (('msg.0289', self._add_record), ('msg.0290', self._copy_record), ('msg.0009', self._rename_record), ('msg.0010', self._delete_record), ('msg.0291', self._toggle_record_enabled), ('msg.0292', self._set_record_group))
        for index, (text, command) in enumerate(record_buttons):
            button_style = 'Danger.TButton' if command == self._delete_record else 'TButton'
            ttk.Button(lb, text=text, command=command, style=button_style).grid(row=index // 3, column=index % 3, padx=2, pady=2, sticky='ew')
        for column in range(3):
            lb.columnconfigure(column, weight=1)
        io_buttons = ttk.Frame(left)
        io_buttons.pack(fill='x', pady=(12, 0))
        ttk.Separator(io_buttons).pack(fill='x', pady=(0, 8))
        ttk.Label(io_buttons, text='msg.0351', style='Subtle.TLabel').pack(anchor='w', pady=(0, 5))
        io_actions = (('msg.0263', self._import_json), ('msg.0262', self._export_json), ('msg.0293', self._import_excel), ('msg.0294', self._export_excel))
        io_grid = ttk.Frame(io_buttons)
        io_grid.pack(fill='x')
        for index, (text, command) in enumerate(io_actions):
            ttk.Button(io_grid, text=text, command=command, style='Toolbar.TButton').grid(row=index // 2, column=index % 2, padx=2, pady=2, sticky='ew')
        for column in range(2):
            io_grid.columnconfigure(column, weight=1, uniform='pcl_io')
        ttk.Label(right, text='msg.0295', style='Section.TLabel').pack(anchor='w')
        tree_frame, self.tree = scrollable_tree(right, columns=('type', 'value', 'path'), show='tree headings')
        self.tree.heading('#0', text='msg.0296')
        self.tree.heading('type', text='msg.0257')
        self.tree.heading('value', text='msg.0379')
        self.tree.heading('path', text='msg.0258')
        self.tree.column('#0', width=200)
        self.tree.column('type', width=80)
        self.tree.column('value', width=220)
        self.tree.column('path', width=220)
        tree_frame.pack(fill='both', expand=True, pady=5)
        self.tree.bind('<Double-1>', self._edit_value)
        self.tree.bind('<Motion>', self._value_column_motion)
        self.tree.bind('<Leave>', lambda _event: self.tree.configure(cursor=''))
        rb = ttk.Frame(right)
        rb.pack(fill='x')
        ttk.Button(rb, text='msg.0298', command=lambda: self._add_list_item(False)).pack(side='left', padx=3)
        ttk.Button(rb, text='msg.0299', command=lambda: self._add_list_item(True)).pack(side='left', padx=3)
        ttk.Button(rb, text='msg.0300', command=self._delete_list_item, style='Danger.TButton').pack(side='left', padx=3)
        ttk.Button(rb, text='msg.0301', command=self._sync_all_records).pack(side='left', padx=3)
        ttk.Button(rb, text='msg.0302', command=lambda: self._save_record(show_message=True), style='Primary.TButton').pack(side='right', padx=3)
        self._sync_all_records(show_message=False)
        self._refresh_records()
        self.transient(parent)

    def _refresh_records(self, select_id: int | None=None) -> None:
        self.records.delete(*self.records.get_children())
        rows = self.db.list_data_records(self.workflow_id)
        if self.record_sort_column:
            rows.sort(key=self._record_sort_key, reverse=self.record_sort_descending)
        for row in rows:
            item = self.records.insert('', 'end', iid=str(row['id']), values=(row['name'], 'msg.0248' if row['enabled'] else 'msg.0309', row['execution_group']), tags=() if row['enabled'] else ('disabled',))
            if row['id'] == select_id:
                self.records.selection_set(item)

    @staticmethod
    def _natural_sort_key(value: Any) -> tuple[tuple[int, Any], ...]:
        """数字を含む名称を、人が期待する順序で比較できるキーへ変換する。"""
        return tuple((0, int(part)) if part.isdigit() else (1, part.casefold()) for part in re.split(r'(\d+)', str(value)))

    def _record_sort_key(self, row: dict[str, Any]) -> Any:
        if self.record_sort_column == 'name':
            return self._natural_sort_key(row['name'])
        if self.record_sort_column == 'enabled':
            return int(row['enabled'])
        if self.record_sort_column == 'group':
            return self._natural_sort_key(row['execution_group'])
        return 0

    def _update_record_headings(self) -> None:
        for column, text_key in self.record_heading_keys.items():
            indicator = ''
            if column == self.record_sort_column:
                indicator = ' ▼' if self.record_sort_descending else ' ▲'
            self.records.heading(column, text=f'{text_key}{indicator}')

    def _sort_records(self, column: str) -> None:
        # 同じ見出しを再度ダブルクリックすると、現在の並び順を反転する。
        if column == self.record_sort_column:
            self.record_sort_descending = not self.record_sort_descending
        else:
            self.record_sort_column = column
            self.record_sort_descending = False
        selection = self.records.selection()
        selected_id = int(selection[0]) if selection else None
        self._update_record_headings()
        self._refresh_records(selected_id)

    def _save_session_limit(self, _event: object=None) -> None:
        try:
            limit = int(self.session_limit.get())
            self.db.set_pcl_session_limit(limit)
        except (TypeError, ValueError):
            self.session_limit.set(str(self.db.get_pcl_session_limit()))

    def _refresh_execution_statuses(self) -> None:
        # 実行処理は別スレッドで DB を更新するため、画面側は定期的に状態だけ同期する。
        return

    def _toggle_record_enabled(self) -> None:
        selection = self.records.selection()
        if not selection:
            messagebox.showinfo('msg.0310', 'msg.0311', parent=self)
            return
        record_id = int(selection[0])
        row = next((row for row in self.db.list_data_records(self.workflow_id) if row['id'] == record_id))
        self.db.set_data_record_enabled(record_id, not row['enabled'])
        self._refresh_records(record_id)

    def _record_double_click(self, event: tk.Event) -> str | None:
        # 実行列は有効/無効を反転し、グループ列は名前入力を開く。
        column = self.records.identify_column(event.x)
        if self.records.identify_region(event.x, event.y) == 'heading':
            column_name = {'#1': 'name', '#2': 'enabled', '#3': 'group'}.get(column)
            if column_name:
                self._sort_records(column_name)
            return 'break'
        if column not in {'#2', '#3'}:
            return None
        item = self.records.identify_row(event.y)
        if not item:
            return None
        self.records.selection_set(item)
        self.records.focus(item)
        if column == '#2':
            self._toggle_record_enabled()
        else:
            self._set_record_group()
        return 'break'

    def _set_record_group(self) -> None:
        selection = self.records.selection()
        if not selection:
            messagebox.showinfo('msg.0310', 'msg.0312', parent=self)
            return
        record_id = int(selection[0])
        row = next((row for row in self.db.list_data_records() if row['id'] == record_id))
        group = simpledialog.askstring('msg.0313', 'msg.0314', initialvalue=row['execution_group'], parent=self)
        if group is None:
            return
        try:
            self.db.set_data_record_group(record_id, group)
        except ValueError as error:
            messagebox.showerror('msg.0315', str(error), parent=self)
            return
        self._refresh_records(record_id)

    def _select_record(self, _event: object=None) -> None:
        selection = self.records.selection()
        if not selection:
            return
        row = next((row for row in self.db.list_data_records(self.workflow_id) if row['id'] == int(selection[0])))
        self.current_id, self.current_name = (row['id'], row['name'])
        self.current_data = normalize_record(self.schema, row['data'])
        self.db.update_data_record(self.current_id, self.current_name, self.current_data)
        self._render()

    def _render(self) -> None:
        had_previous_view = bool(self.view_identity)
        expanded = {identity for item, identity in self.view_identity.items() if self.tree.exists(item) and bool(self.tree.item(item, 'open'))}
        selected_items = self.tree.selection()
        selected_identity = self.view_identity.get(selected_items[0]) if selected_items else None
        self.tree.delete(*self.tree.get_children())
        self.meta.clear()
        self.view_identity.clear()

        def open_state(identity: str, default: bool) -> bool:
            if identity in self.force_open_identities:
                return True
            return identity in expanded if had_previous_view else default

        def show(parent_item: str, node: dict[str, Any], value: Any, parent_value: Any, key: Any, path: str, identity: str, title: str | None=None, item_container: list[Any] | None=None, item_index: int | None=None) -> None:
            kind = node['type']
            # 直接編集できる値だけに鉛筆記号を付け、構造行との違いを明確にする。
            display = '' if kind in ('object', 'list') else f'✎ {value}'
            default_open = kind in ('object', 'list') and (not (kind == 'list' and (not value)))
            item = self.tree.insert(parent_item, 'end', text=title or node['name'], values=(kind, display, path), open=open_state(identity, default_open))
            self.meta[item] = {'node': node, 'value': value, 'parent': parent_value, 'key': key, 'list': item_container, 'index': item_index, 'identity': identity}
            self.view_identity[item] = identity
            if kind == 'object':
                for child in node.get('children', []):
                    child_path = f'{path}.{child['name']}' if path else child['name']
                    show(item, child, value.get(child['name'], default_value(child)), value, child['name'], child_path, f'{identity}/{child['name']}')
            elif kind == 'list':
                for index, child_value in enumerate(value):
                    wrapper = {'name': f'[{index + 1}]', 'type': 'object', 'children': node.get('children', [])}
                    show(item, wrapper, child_value, value, index, path, f'{identity}[{index}]', f'[{index + 1}]', value, index)
                if not value:
                    placeholder_identity = f'{identity}/__empty__'
                    placeholder = self.tree.insert(item, 'end', text='msg.0316', values=('msg.0048', 'msg.0317', path), open=open_state(placeholder_identity, True))
                    self.view_identity[placeholder] = placeholder_identity

                    def show_expected(parent: str, children: list[dict[str, Any]], expected_prefix: str, parent_identity: str) -> None:
                        for child in children:
                            child_path = f'{expected_prefix}.{child['name']}'
                            child_identity = f'{parent_identity}/{child['name']}'
                            expected = self.tree.insert(parent, 'end', text=child['name'], values=(child['type'], 'msg.0318', child_path), open=open_state(child_identity, True))
                            self.view_identity[expected] = child_identity
                            if child['type'] in ('object', 'list'):
                                show_expected(expected, child.get('children', []), child_path, child_identity)
                    show_expected(placeholder, node.get('children', []), path, placeholder_identity)
        root = {'name': 'msg.0279', 'type': 'object', 'children': self.schema.get('children', [])}
        show('', root, self.current_data, None, None, '', 'root')
        identity_to_select = self.force_select_identity or selected_identity
        if identity_to_select:
            for item, identity in self.view_identity.items():
                if identity == identity_to_select:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    break
        self.force_open_identities.clear()
        self.force_select_identity = None

    def _add_record(self) -> None:
        entered_name = simpledialog.askstring('msg.0289', 'msg.0319', parent=self)
        if entered_name is None:
            return
        name = entered_name.strip() or tr('msg.0320')
        data = {child['name']: default_value(child) for child in self.schema.get('children', [])}
        record_id = self.db.add_data_record(self.workflow_id, name, data)
        self._refresh_records(record_id)
        self._select_record()

    def _rename_record(self) -> None:
        if self.current_id is None:
            return
        name = simpledialog.askstring('msg.0044', 'msg.0319', initialvalue=self.current_name, parent=self)
        if name:
            self.current_name = name
            self._save_record()
            self._refresh_records(self.current_id)

    def _unique_name(self, base: str) -> str:
        names = {row['name'] for row in self.db.list_data_records(self.workflow_id)}
        if base not in names:
            return base
        number = 2
        while f'{base} ({number})' in names:
            number += 1
        return f'{base} ({number})'

    def _copy_record(self) -> None:
        if self.current_id is None:
            messagebox.showinfo('msg.0310', 'msg.0321', parent=self)
            return
        self._save_record()
        name = self._unique_name(f'{self.current_name}{tr("msg.0322")}')
        record_id = self.db.add_data_record(self.workflow_id, name, copy.deepcopy(self.current_data))
        self._refresh_records(record_id)
        self._select_record()

    def _delete_record(self) -> None:
        if self.current_id is not None and messagebox.askyesno('msg.0046', 'msg.0323', parent=self):
            self.db.delete_data_record(self.workflow_id, self.current_id)
            self.current_id = None
            self.tree.delete(*self.tree.get_children())
            self._refresh_records()

    def _value_column_motion(self, event: tk.Event) -> None:
        item = self.tree.identify_row(event.y)
        meta = self.meta.get(item)
        editable = self.tree.identify_region(event.x, event.y) == 'cell' and self.tree.identify_column(event.x) == '#2' and meta is not None and meta['node']['type'] not in ('object', 'list')
        self.tree.configure(cursor='hand2' if editable else '')

    def _edit_value(self, event: tk.Event) -> None:
        if self.tree.identify_column(event.x) != '#2':
            return
        item = self.tree.identify_row(event.y)
        meta = self.meta.get(item)
        if not meta or meta['node']['type'] in ('object', 'list'):
            return
        old = str(meta['value'])
        value = simpledialog.askstring('msg.0324', meta['node']['name'], initialvalue=old, parent=self)
        if value is None:
            return
        kind = meta['node']['type']
        try:
            yes_words = {tr_language('msg.0037', language).lower() for language in SUPPORTED_LANGUAGES}
            parsed: Any = float(value) if kind == 'number' and '.' in value else int(value) if kind == 'number' else value.lower() in {'true', '1', *yes_words} if kind == 'boolean' else value
        except ValueError:
            messagebox.showerror('msg.0159', 'msg.0325', parent=self)
            return
        meta['parent'][meta['key']] = parsed
        self._render()
        self._save_record()

    def _selected_list(self) -> dict[str, Any] | None:
        selection = self.tree.selection()
        item = selection[0] if selection else ''
        while item:
            meta = self.meta.get(item)
            if meta and meta['node']['type'] == 'list':
                return meta
            item = self.tree.parent(item)
        return None

    def _add_list_item(self, complete: bool=False) -> None:
        meta = self._selected_list()
        if not meta or meta['node']['type'] != 'list':
            messagebox.showinfo('msg.0281', 'msg.0326', parent=self)
            return
        new_index = len(meta['value'])
        self.force_open_identities.update({meta['identity'], f'{meta['identity']}[{new_index}]'})
        self.force_select_identity = f'{meta['identity']}[{new_index}]'
        item = new_list_item(meta['node'], populate_nested_lists=complete)
        meta['value'].append(item)
        self._render()
        self._save_record()

    def _delete_list_item(self) -> None:
        selection = self.tree.selection()
        meta = self.meta.get(selection[0]) if selection else None
        if not meta or meta.get('list') is None:
            messagebox.showinfo('msg.0327', 'msg.0328', parent=self)
            return
        del meta['list'][meta['index']]
        self._render()
        self._save_record()

    def _save_record(self, show_message: bool=False) -> bool:
        if self.current_id is None:
            if show_message:
                messagebox.showinfo('msg.0048', 'msg.0443', parent=self)
            return False
        try:
            self.db.update_data_record(self.current_id, self.current_name, self.current_data)
        except Exception as error:
            if show_message:
                messagebox.showerror('msg.0159', f'msg.0442{error}', parent=self)
            return False
        if show_message:
            messagebox.showinfo('msg.0306', 'msg.0441', parent=self)
        return True

    def _sync_all_records(self, show_message: bool=True) -> None:
        records = self.db.list_data_records(self.workflow_id)
        for record in records:
            synchronized = normalize_record(self.schema, record['data'])
            self.db.update_data_record(record['id'], record['name'], synchronized)
        current_id = self.current_id
        if current_id is not None:
            self._refresh_records(current_id)
            self._select_record()
        if show_message:
            messagebox.showinfo('msg.0329', f'msg.0330{len(records)}msg.0331', parent=self)

    def _export_json(self) -> None:
        self._save_record()
        records = self.db.list_data_records(self.workflow_id)
        if not records:
            messagebox.showinfo('msg.0332', 'msg.0333', parent=self)
            return
        path = filedialog.asksaveasfilename(parent=self, defaultextension='.json', filetypes=(('Json', '*.json'),))
        if not path:
            return
        payload = {'version': 1, 'type': 'web-flow-data', 'records': [{'name': row['name'], 'enabled': row['enabled'], 'execution_group': row['execution_group'], 'data': row['data']} for row in records]}
        Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        messagebox.showinfo('msg.0269', f'msg.0334{len(records)}msg.0335{path}', parent=self)

    def _import_json(self) -> None:
        path = filedialog.askopenfilename(parent=self, filetypes=(('Json', '*.json'), ('msg.0054', '*.*')))
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding='utf-8-sig'))
            records = payload.get('records') if isinstance(payload, dict) else None
            if not isinstance(records, list):
                raise ValueError('msg.0336')
            checked: list[tuple[str, bool, str, dict[str, Any]]] = []
            for index, record in enumerate(records, 1):
                if not isinstance(record, dict) or not isinstance(record.get('name'), str) or (not isinstance(record.get('data'), dict)):
                    raise ValueError(f'msg.0103{index}msg.0337')
                enabled = record.get('enabled', True)
                if not isinstance(enabled, bool):
                    raise ValueError(f'msg.0103{index}msg.0338')
                group = str(record.get('execution_group', '1')).strip()
                if not group:
                    raise ValueError(f'msg.0103{index}msg.0339')
                checked.append((record['name'].strip() or f'Data {index}', enabled, group, normalize_record(self.schema, record['data'])))
        except (OSError, json.JSONDecodeError, ValueError) as error:
            messagebox.showerror('msg.0057', str(error), parent=self)
            return
        for name, enabled, group, data in checked:
            record_id = self.db.add_data_record(self.workflow_id, self._unique_name(name), data)
            self.db.set_data_record_enabled(record_id, enabled)
            self.db.set_data_record_group(record_id, group)
        self._refresh_records()
        messagebox.showinfo('msg.0274', f'msg.0340{len(checked)}msg.0341', parent=self)

    def _export_excel(self) -> None:
        self._save_record()
        records = self.db.list_data_records(self.workflow_id)
        if not records:
            messagebox.showinfo('msg.0332', 'msg.0333', parent=self)
            return
        path = filedialog.asksaveasfilename(parent=self, defaultextension='.xlsx', filetypes=(('Excel', '*.xlsx'),))
        if not path:
            return
        try:
            row_count = write_records_excel(path, self.schema, records)
        except ImportError:
            messagebox.showerror('msg.0342', 'msg.0343', parent=self)
            return
        except OSError as error:
            messagebox.showerror('msg.0344', str(error), parent=self)
            return
        messagebox.showinfo('msg.0269', f'msg.0345{row_count}msg.0346{path}', parent=self)

    def _import_excel(self) -> None:
        # Excel 読込は追加ではなく完全復元である。確認後に一つのトランザクションで置換する。
        path = filedialog.askopenfilename(parent=self, filetypes=(('Excel', '*.xlsx'), ('msg.0054', '*.*')))
        if not path:
            return
        try:
            records = read_records_excel(path, self.schema)
        except ImportError:
            messagebox.showerror('msg.0342', 'msg.0343', parent=self)
            return
        except (OSError, ValueError) as error:
            messagebox.showerror('msg.0057', str(error), parent=self)
            return
        if not messagebox.askyesno('msg.0347', f'msg.0348{len(records)}msg.0349', parent=self):
            return
        self.current_id = None
        self.current_name = ''
        self.current_data = {}
        self.db.replace_data_records(records)
        self._refresh_records()
        self.tree.delete(*self.tree.get_children())
        messagebox.showinfo('msg.0274', f'msg.0350{len(records)}msg.0331', parent=self)
